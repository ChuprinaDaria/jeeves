"""Recalculate Excel formulas using LibreOffice headless.

Usage: python scripts/recalc.py <excel_file> [timeout_seconds]
Returns JSON with error details.
"""
import json
import subprocess
import sys
import tempfile
from pathlib import Path


def recalculate(filepath: str, timeout: int = 60) -> dict:
    filepath = Path(filepath).resolve()
    if not filepath.exists():
        return {'status': 'error', 'message': f'File not found: {filepath}'}

    # Use a temp dir for LibreOffice profile to avoid locking issues
    with tempfile.TemporaryDirectory(prefix='lo_recalc_') as profile_dir:
        cmd = [
            'libreoffice',
            '--headless',
            '--calc',
            '--norestore',
            f'-env:UserInstallation=file://{profile_dir}',
            '--convert-to', 'xlsx',
            '--outdir', str(filepath.parent),
            str(filepath),
        ]

        try:
            result = subprocess.run(
                cmd, capture_output=True, text=True, timeout=timeout)
        except subprocess.TimeoutExpired:
            return {'status': 'error', 'message': f'Timeout after {timeout}s'}
        except FileNotFoundError:
            return {'status': 'error', 'message': 'LibreOffice not installed'}

        if result.returncode != 0:
            return {
                'status': 'error',
                'message': f'LibreOffice error: {result.stderr.strip()}',
            }

    # Scan for formula errors
    return scan_errors(filepath)


def scan_errors(filepath: Path) -> dict:
    try:
        from openpyxl import load_workbook
    except ImportError:
        return {'status': 'error', 'message': 'openpyxl not installed'}

    # Load with formulas to count them
    wb_formulas = load_workbook(str(filepath), data_only=False)
    total_formulas = 0
    for sheet in wb_formulas.sheetnames:
        ws = wb_formulas[sheet]
        for row in ws.iter_rows():
            for cell in row:
                if cell.value and isinstance(cell.value, str) and cell.value.startswith('='):
                    total_formulas += 1
    wb_formulas.close()

    # Load with cached values to find errors
    wb = load_workbook(str(filepath), data_only=True)
    error_types = ('#REF!', '#DIV/0!', '#VALUE!', '#N/A', '#NAME?', '#NULL!', '#NUM!')
    error_summary = {}
    total_errors = 0

    for sheet_name in wb.sheetnames:
        sheet = wb[sheet_name]
        for row in sheet.iter_rows():
            for cell in row:
                if cell.value in error_types:
                    total_errors += 1
                    err_type = str(cell.value)
                    if err_type not in error_summary:
                        error_summary[err_type] = {'count': 0, 'locations': []}
                    error_summary[err_type]['count'] += 1
                    error_summary[err_type]['locations'].append(
                        f'{sheet_name}!{cell.coordinate}')

    wb.close()

    return {
        'status': 'errors_found' if total_errors > 0 else 'success',
        'total_errors': total_errors,
        'total_formulas': total_formulas,
        'error_summary': error_summary if error_summary else None,
    }


if __name__ == '__main__':
    if len(sys.argv) < 2:
        print(json.dumps({'status': 'error', 'message': 'Usage: python recalc.py <file> [timeout]'}))
        sys.exit(1)

    fpath = sys.argv[1]
    tout = int(sys.argv[2]) if len(sys.argv) > 2 else 60
    result = recalculate(fpath, tout)
    print(json.dumps(result, indent=2))
    sys.exit(0 if result['status'] == 'success' else 1)
