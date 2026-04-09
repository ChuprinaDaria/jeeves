"""
mcp-xlsx -- FastMCP server for creating professional Excel spreadsheets.

Provides:
- create_spreadsheet(client_id, filename, sheets) tool for generating .xlsx files
  with formulas, formatting, and automatic recalculation.
"""

from __future__ import annotations

import importlib.util
import json
import logging
import os
from pathlib import Path

from mcp_servers.common.django_setup import setup

setup()

from asgiref.sync import sync_to_async  # noqa: E402
from fastmcp import FastMCP  # noqa: E402

logger = logging.getLogger(__name__)

mcp = FastMCP(
    "mcp-xlsx",
    description="Nexelin XLSX processor. "
    "Use the `create_spreadsheet` tool to generate professional Excel files "
    "with formulas, formatting, and automatic recalculation.",
)

XLSX_OUTPUT_DIR = Path(os.environ.get(
    'XLSX_OUTPUT_DIR',
    Path(__file__).resolve().parent.parent.parent / 'media' / 'xlsx',
))

RECALC_SCRIPT = Path(__file__).resolve().parent.parent.parent / 'scripts' / 'recalc.py'


# ---------------------------------------------------------------------------
# Sync helpers (openpyxl + Django ORM are synchronous)
# ---------------------------------------------------------------------------

def _create_workbook_sync(
    client_id: int,
    filename: str = 'report.xlsx',
    sheets: list[dict] | None = None,
) -> dict:
    """Create a professional Excel workbook and run formula recalculation."""
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, NamedStyle, PatternFill
    from openpyxl.utils import get_column_letter

    output_dir = XLSX_OUTPUT_DIR / str(client_id)
    output_dir.mkdir(parents=True, exist_ok=True)

    if not sheets:
        sheets = [{'name': 'Sheet1', 'headers': [], 'rows': []}]

    wb = Workbook()

    # --- Styles ---
    header_font = Font(name='Arial', bold=True, size=11)
    header_fill = PatternFill('solid', fgColor='E2EFDA')
    header_alignment = Alignment(horizontal='center')

    data_font = Font(name='Arial', size=11, color='0000FF')        # blue for inputs
    formula_font = Font(name='Arial', size=11, color='000000')      # black for formulas
    crossref_font = Font(name='Arial', size=11, color='006100')     # green for cross-sheet

    errors: list[str] = []
    sheet_names: list[str] = []

    for i, sheet_spec in enumerate(sheets):
        sheet_name = sheet_spec.get('name', f'Sheet{i + 1}')
        sheet_names.append(sheet_name)

        if i == 0:
            ws = wb.active
            ws.title = sheet_name
        else:
            ws = wb.create_sheet(sheet_name)

        # -- Headers --
        headers = sheet_spec.get('headers', [])
        if headers:
            ws.append(headers)
            for col_idx in range(1, len(headers) + 1):
                cell = ws.cell(row=1, column=col_idx)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = header_alignment

        # -- Data rows --
        for row_data in sheet_spec.get('rows', []):
            ws.append(row_data)

        # -- Apply input font to data cells --
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row, max_col=ws.max_column):
            for cell in row:
                if cell.value is not None and not (
                    isinstance(cell.value, str) and cell.value.startswith('=')
                ):
                    cell.font = data_font

        # -- Formulas --
        for cell_ref, formula in sheet_spec.get('formulas', {}).items():
            ws[cell_ref] = formula
            # Color: green for cross-sheet references, black otherwise
            if '!' in formula:
                ws[cell_ref].font = crossref_font
            else:
                ws[cell_ref].font = formula_font

        # -- Column widths --
        for col_letter, width in sheet_spec.get('column_widths', {}).items():
            ws.column_dimensions[col_letter].width = width

        # -- Auto-width fallback for columns without explicit width --
        explicit_widths = set(sheet_spec.get('column_widths', {}).keys())
        if headers:
            for col_idx, header in enumerate(headers, 1):
                col_letter = get_column_letter(col_idx)
                if col_letter not in explicit_widths:
                    ws.column_dimensions[col_letter].width = max(len(str(header)) + 4, 12)

    # -- Number formatting pass --
    for ws in wb.worksheets:
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row, max_col=ws.max_column):
            for cell in row:
                val = cell.value
                if val is None:
                    continue
                # Formulas: apply format based on heuristics from the formula text
                if isinstance(val, str) and val.startswith('='):
                    val_lower = val.lower()
                    if any(kw in val_lower for kw in ('%', 'percent', 'margin', 'rate')):
                        cell.number_format = '0.0%'
                # Numeric values: no special global format -- let caller control via formulas

    filepath = output_dir / filename
    wb.save(str(filepath))
    wb.close()

    # -- Recalculate formulas via LibreOffice --
    recalc_result = _run_recalc(str(filepath))
    if recalc_result.get('status') == 'error':
        errors.append(f"Recalc: {recalc_result.get('message', 'unknown error')}")
    elif recalc_result.get('status') == 'errors_found':
        summary = recalc_result.get('error_summary', {})
        for err_type, info in (summary or {}).items():
            locations = ', '.join(info.get('locations', [])[:5])
            errors.append(f"Formula {err_type} at {locations}")

    download_url = f'/media/xlsx/{client_id}/{filename}'

    return {
        'status': 'created' if not errors else 'created_with_warnings',
        'filepath': str(filepath),
        'filename': filename,
        'download_url': download_url,
        'sheets': sheet_names,
        'errors': errors if errors else None,
    }


def _run_recalc(filepath: str) -> dict:
    """Run formula recalculation via scripts/recalc.py."""
    if not RECALC_SCRIPT.exists():
        return {'status': 'error', 'message': 'recalc.py script not found'}

    try:
        spec = importlib.util.spec_from_file_location('recalc', str(RECALC_SCRIPT))
        recalc_mod = importlib.util.module_from_spec(spec)
        spec.loader.exec_module(recalc_mod)
        return recalc_mod.recalculate(filepath)
    except Exception as e:
        logger.error(f'Recalc failed for {filepath}: {e}')
        return {'status': 'error', 'message': str(e)}


# ---------------------------------------------------------------------------
# MCP tool
# ---------------------------------------------------------------------------

@mcp.tool()
async def create_spreadsheet(
    client_id: int,
    filename: str = 'report.xlsx',
    sheets: list[dict] | None = None,
) -> str:
    """Create a professional Excel spreadsheet with data, formulas, and formatting.

    Use this tool when you need to generate .xlsx files for clients.
    Supports multiple sheets, Excel formulas, column widths, and automatic
    formula recalculation via LibreOffice.

    Formulas must be real Excel formulas (e.g. "=SUM(B2:B9)"), NOT hardcoded values.
    Cross-sheet references use standard Excel syntax (e.g. "=Sheet1!B2*0.1").

    Args:
        client_id: Numeric ID of the Nexelin client (used for file isolation).
        filename: Output filename (default "report.xlsx").
        sheets: List of sheet specifications, each a dict with:
            - name: sheet tab name
            - headers: list of column header strings
            - rows: list of row data (each row is a list of values)
            - formulas: dict mapping cell references to Excel formulas
              e.g. {"B10": "=SUM(B2:B9)", "C5": "=B5/B10"}
            - column_widths: dict mapping column letters to widths
              e.g. {"A": 30, "B": 15}
    """
    result = await sync_to_async(_create_workbook_sync)(client_id, filename, sheets)
    return json.dumps(result, ensure_ascii=False)


# ---------------------------------------------------------------------------
# Entrypoint
# ---------------------------------------------------------------------------

if __name__ == "__main__":
    mcp.run(transport="stdio")
