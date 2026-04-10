"""Builtin XLSX processor — create, read, edit, and analyze Excel files."""
import logging
import os
from pathlib import Path

from asgiref.sync import sync_to_async

logger = logging.getLogger(__name__)

# Media dir for generated files
XLSX_OUTPUT_DIR = Path(os.environ.get(
    'XLSX_OUTPUT_DIR',
    Path(__file__).resolve().parent.parent.parent.parent / 'media' / 'xlsx'))


def _validate_path(filepath, output_dir):
    """Ensure filepath is within allowed output directory (prevent path traversal)."""
    resolved = Path(filepath).resolve()
    allowed = Path(output_dir).resolve()
    if not resolved.is_relative_to(allowed):
        return None, {'error': 'Access denied: path outside allowed directory'}
    return resolved, None


async def xlsx_processor(connection, tool_name, action='create', **kwargs):
    """Process XLSX operations.

    Actions:
        create  — create new workbook from data/formulas/formatting spec
        read    — read and return data from existing file
        edit    — modify existing file preserving formulas
        recalc  — recalculate formulas via LibreOffice
    """
    client = connection.client
    output_dir = XLSX_OUTPUT_DIR / str(client.pk)
    output_dir.mkdir(parents=True, exist_ok=True)

    handlers = {
        'create': _create_workbook,
        'read': _read_workbook,
        'edit': _edit_workbook,
        'recalc': _recalculate,
    }

    handler = handlers.get(action)
    if not handler:
        return {'error': f'Unknown action: {action}. Use: {", ".join(handlers)}'}

    try:
        return await sync_to_async(handler)(output_dir=output_dir, **kwargs)
    except Exception as e:
        logger.error(f'XLSX processor error for client {client.pk}: {e}')
        return {'error': str(e)}


def _create_workbook(output_dir, filename='report.xlsx', sheets=None, **kwargs):
    """Create a new Excel workbook.

    Args:
        filename: output filename
        sheets: list of sheet specs, each with:
            - name: sheet name
            - headers: list of column headers
            - rows: list of row data (lists)
            - formulas: dict of {cell: formula} e.g. {"B10": "=SUM(B2:B9)"}
            - column_widths: dict of {col_letter: width}
    """
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill

    wb = Workbook()

    if not sheets:
        sheets = [{'name': 'Sheet1', 'headers': [], 'rows': []}]

    for i, sheet_spec in enumerate(sheets):
        if i == 0:
            ws = wb.active
            ws.title = sheet_spec.get('name', 'Sheet1')
        else:
            ws = wb.create_sheet(sheet_spec.get('name', f'Sheet{i + 1}'))

        # Headers
        headers = sheet_spec.get('headers', [])
        if headers:
            ws.append(headers)
            for col_idx, _ in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col_idx)
                cell.font = Font(name='Arial', bold=True, size=11)
                cell.fill = PatternFill('solid', fgColor='E2EFDA')
                cell.alignment = Alignment(horizontal='center')

        # Data rows
        for row_data in sheet_spec.get('rows', []):
            ws.append(row_data)

        # Formulas
        for cell_ref, formula in sheet_spec.get('formulas', {}).items():
            ws[cell_ref] = formula
            ws[cell_ref].font = Font(name='Arial', color='000000')

        # Column widths
        for col_letter, width in sheet_spec.get('column_widths', {}).items():
            ws.column_dimensions[col_letter].width = width

        # Default font for data cells
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row,
                                max_col=ws.max_column):
            for cell in row:
                if cell.font == Font():
                    cell.font = Font(name='Arial', size=11)

    filepath = output_dir / filename
    wb.save(str(filepath))
    wb.close()

    return {
        'status': 'created',
        'filepath': str(filepath),
        'filename': filename,
        'sheets': [s.get('name', f'Sheet{i+1}') for i, s in enumerate(sheets)],
    }


def _read_workbook(filepath=None, sheet_name=None, output_dir=None, **kwargs):
    """Read Excel file and return structured data."""
    import pandas as pd

    if not filepath:
        return {'error': 'filepath is required'}

    path, err = _validate_path(filepath, output_dir)
    if err:
        return err
    if not path.exists():
        return {'error': f'File not found: {filepath}'}

    try:
        if sheet_name:
            df = pd.read_excel(str(path), sheet_name=sheet_name)
            return {
                'status': 'ok',
                'sheet': sheet_name,
                'columns': list(df.columns),
                'row_count': len(df),
                'data': df.head(100).to_dict(orient='records'),
                'dtypes': {col: str(dt) for col, dt in df.dtypes.items()},
            }
        else:
            all_sheets = pd.read_excel(str(path), sheet_name=None)
            result = {'status': 'ok', 'sheets': {}}
            for name, df in all_sheets.items():
                result['sheets'][name] = {
                    'columns': list(df.columns),
                    'row_count': len(df),
                    'data': df.head(50).to_dict(orient='records'),
                }
            return result
    except Exception as e:
        return {'error': f'Failed to read: {e}'}


def _edit_workbook(filepath=None, operations=None, output_dir=None, **kwargs):
    """Edit existing Excel file preserving formulas and formatting.

    Args:
        filepath: path to existing file
        operations: list of operations:
            - {type: 'set_cell', sheet: 'Sheet1', cell: 'A1', value: 'New Value'}
            - {type: 'set_formula', sheet: 'Sheet1', cell: 'B10', formula: '=SUM(B2:B9)'}
            - {type: 'insert_row', sheet: 'Sheet1', row: 2}
            - {type: 'delete_row', sheet: 'Sheet1', row: 3}
            - {type: 'add_sheet', name: 'NewSheet'}
    """
    from openpyxl import load_workbook

    if not filepath:
        return {'error': 'filepath is required'}

    path, err = _validate_path(filepath, output_dir)
    if err:
        return err
    if not path.exists():
        return {'error': f'File not found: {filepath}'}

    wb = load_workbook(str(path))
    applied = []

    for op in (operations or []):
        op_type = op.get('type')
        sheet_name = op.get('sheet')
        ws = wb[sheet_name] if sheet_name and sheet_name in wb.sheetnames else wb.active

        if op_type == 'set_cell':
            ws[op['cell']] = op['value']
            applied.append(f"set {sheet_name}!{op['cell']}")
        elif op_type == 'set_formula':
            ws[op['cell']] = op['formula']
            applied.append(f"formula {sheet_name}!{op['cell']}")
        elif op_type == 'insert_row':
            ws.insert_rows(op['row'])
            applied.append(f"insert row {op['row']} in {sheet_name}")
        elif op_type == 'delete_row':
            ws.delete_rows(op['row'])
            applied.append(f"delete row {op['row']} in {sheet_name}")
        elif op_type == 'add_sheet':
            wb.create_sheet(op['name'])
            applied.append(f"add sheet {op['name']}")

    wb.save(str(path))
    wb.close()

    return {
        'status': 'edited',
        'filepath': str(path),
        'operations_applied': applied,
    }


def _recalculate(filepath=None, output_dir=None, **kwargs):
    """Recalculate formulas using LibreOffice."""
    if not filepath:
        return {'error': 'filepath is required'}

    path, err = _validate_path(filepath, output_dir)
    if err:
        return err

    import importlib.util
    recalc_path = Path(__file__).resolve().parent.parent.parent.parent / 'scripts' / 'recalc.py'

    if not recalc_path.exists():
        return {'error': 'recalc.py script not found'}

    spec = importlib.util.spec_from_file_location('recalc', str(recalc_path))
    recalc_mod = importlib.util.module_from_spec(spec)
    spec.loader.exec_module(recalc_mod)

    return recalc_mod.recalculate(str(path))
